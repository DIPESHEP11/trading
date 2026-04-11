from io import BytesIO
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from django.http import HttpResponse
from openpyxl import Workbook
from apps.crm.models import (
    Lead, Customer, LeadActivity, LeadFormSchema, LeadSource,
    CustomLeadStatus, CustomLeadSource, StatusFlowAction, LeadAssignmentConfig,
)
from apps.crm.constants import (
    DEFAULT_LEAD_FIELDS, LEAD_CORE_KEYS, CUSTOM_DATA_KEYS, DISPLAY_ONLY_KEYS,
)
from apps.crm.lead_form_utils import (
    merge_default_fields, effective_form_fields, validate_required_from_schema,
    validate_phone_value,
)
from apps.crm.assignment import auto_assign_lead_from_bulk_defaults
from apps.core.responses import success_response, error_response
from apps.core.permissions import IsStaffOrAbove, IsTenantAdmin
from apps.config.utils import log_module_history
from .serializers import (
    LeadSerializer, LeadListSerializer, CustomerSerializer,
    LeadActivitySerializer, LeadFormSchemaSerializer,
    CustomLeadStatusSerializer, CustomLeadSourceSerializer,
    StatusFlowActionSerializer,
)
from .assignment_serializers import LeadAssignmentConfigSerializer


def _crm_tenant():
    from django.db import connection
    return getattr(connection, 'tenant', None)


def _auto_link_customer(lead):
    """Find or create a Customer matching the lead's email/phone, then link it."""
    if not lead.email and not lead.phone:
        return
    customer = None
    if lead.email:
        customer = Customer.objects.filter(email=lead.email).first()
    if not customer and lead.phone:
        customer = Customer.objects.filter(phone=lead.phone).first()
    if customer:
        # Fill any blank fields on the existing customer from lead data
        changed = False
        for attr in ('email', 'phone', 'company'):
            if not getattr(customer, attr) and getattr(lead, attr):
                setattr(customer, attr, getattr(lead, attr))
                changed = True
        if changed:
            customer.save()
    else:
        parts = lead.name.split(None, 1) if lead.name else ['Unknown']
        customer = Customer.objects.create(
            first_name=parts[0],
            last_name=parts[1] if len(parts) > 1 else '',
            email=lead.email or '',
            phone=lead.phone or '',
            company=lead.company or '',
        )
    lead.customer = customer
    lead.save(update_fields=['customer'])


class LeadListCreateView(generics.ListCreateAPIView):
    """GET POST /api/v1/crm/leads/"""
    permission_classes = [IsStaffOrAbove]

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        try:
            s, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
            ctx['lead_form_schema'] = s
        except Exception:
            ctx['lead_form_schema'] = None
        ctx['tenant'] = _crm_tenant()
        return ctx

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return LeadListSerializer
        return LeadSerializer

    def get_queryset(self):
        qs = Lead.objects.all()
        status_filter = self.request.query_params.get('status')
        source = self.request.query_params.get('source')
        assigned = self.request.query_params.get('assigned_to')
        search = self.request.query_params.get('search')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if source:
            qs = qs.filter(source=source)
        if assigned:
            qs = qs.filter(assigned_to=assigned)
        if search:
            qs = qs.filter(name__icontains=search) | qs.filter(phone__icontains=search) | qs.filter(email__icontains=search)
        return qs.select_related('assigned_to', 'customer')

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        data = LeadListSerializer(qs, many=True).data
        return success_response(data={'leads': data, 'count': len(data)})

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        lead = serializer.save()
        _auto_link_customer(lead)
        auto_assign_lead_from_bulk_defaults(lead)
        LeadActivity.objects.create(
            lead=lead, activity_type='status_change',
            description=f'Lead created with status: {lead.status}',
            performed_by=request.user
        )
        log_module_history(request, 'crm', 'create', f'Lead created: {lead.name}', 'lead', str(lead.pk))
        return success_response(data=LeadSerializer(lead).data,
                                message='Lead created.', http_status=status.HTTP_201_CREATED)


def _execute_status_flow(lead, user):
    """
    Check if there's a StatusFlowAction for the lead's current status.
    If found, execute the corresponding cross-module action.
    Returns a dict with flow_result info (or None if no flow).
    """
    try:
        flow = StatusFlowAction.objects.get(status_key=lead.status, is_active=True)
    except StatusFlowAction.DoesNotExist:
        return None

    result = {
        'status_key': flow.status_key,
        'target_module': flow.target_module,
        'action': flow.action,
        'executed': False,
        'message': '',
    }

    if flow.target_module == 'none' or flow.action == 'notify_only':
        result['executed'] = True
        result['message'] = f'Status changed to "{flow.status_key}" (CRM only).'
        return result

    if flow.action == 'create_order' and flow.target_module == 'orders':
        from apps.orders.models import Order, OrderStatusHistory
        existing = Order.objects.filter(external_id=f'lead-{lead.id}').first()
        if existing:
            result['executed'] = True
            result['message'] = f'Order #{existing.order_number} already exists for this lead.'
            return result

        # ── Collect ALL lead data ──────────────────────────────────────────
        custom = lead.custom_data or {}
        # Map common custom field names to order shipping fields
        address = custom.get('address', '') or ''
        city = custom.get('city', '') or ''
        state = custom.get('state', '') or ''
        pincode = custom.get('pincode', '') or custom.get('zip', '') or ''
        shipping_address = address

        # Build structured notes that carry ALL lead info
        lead_data_summary = []
        for k, v in custom.items():
            if v and k not in ('address', 'city', 'state', 'pincode', 'zip'):
                lead_data_summary.append(f'{k}: {v}')
        custom_notes_str = '\n'.join(lead_data_summary) if lead_data_summary else ''
        full_notes = (
            f'Auto-created from lead #{lead.id} — {lead.name}'
            + (f' | Company: {lead.company}' if lead.company else '')
            + (f'\nLead notes: {lead.notes}' if lead.notes else '')
            + (f'\nCustom fields:\n{custom_notes_str}' if custom_notes_str else '')
        )

        order = Order.objects.create(
            source=lead.source or 'manual',
            customer=lead.customer,
            shipping_name=lead.name,
            shipping_phone=lead.phone or '',
            shipping_address=shipping_address,
            shipping_city=city,
            shipping_state=state,
            shipping_pincode=pincode,
            notes=full_notes,
            external_id=f'lead-{lead.id}',
            assigned_to=lead.assigned_to or user,
        )
        OrderStatusHistory.objects.create(
            order=order, from_status='', to_status=order.status,
            changed_by=user, note=f'Auto-created from CRM lead #{lead.id}')
        LeadActivity.objects.create(
            lead=lead, activity_type='status_change',
            description=f'Flow: Order #{order.order_number} created in Orders module.',
            performed_by=user)
        result['executed'] = True
        result['message'] = f'Order #{order.order_number} created.'
        return result

    if flow.action == 'send_to_warehouse' and flow.target_module == 'warehouse':
        from apps.orders.models import Order
        from apps.warehouses.models import InventoryApproval
        order = Order.objects.filter(external_id=f'lead-{lead.id}').select_related('customer').prefetch_related('items__product').first()
        if order:
            last_apr = InventoryApproval.objects.order_by('-id').first()
            num = (last_apr.id + 1) if last_apr else 1
            req_number = f'INV-APR-{num:05d}'
            order_items = []
            for oi in order.items.all():
                order_items.append({
                    'product_id': oi.product_id,
                    'product_name': oi.product.name,
                    'quantity': float(oi.quantity),
                    'notes': f'Lead: {lead.name} — ₹{oi.unit_price}/unit',
                })
            # Collect complete upstream context
            extra_data = {
                'lead_id': lead.id,
                'lead_name': lead.name,
                'lead_phone': lead.phone,
                'lead_email': lead.email,
                'lead_company': lead.company,
                'lead_source': lead.source,
                'lead_status': lead.status,
                'lead_notes': lead.notes,
                'lead_custom_data': lead.custom_data or {},
                'order_number': order.order_number,
                'order_shipping_name': order.shipping_name,
                'order_shipping_phone': order.shipping_phone,
                'order_shipping_address': order.shipping_address,
                'order_total': str(order.total_amount),
                'order_notes': order.notes,
            }
            InventoryApproval.objects.create(
                request_number=req_number,
                source_module='crm',
                source_reference=order.order_number,
                requested_action='stock_out',
                next_module='dispatch',
                notes=f'Auto from CRM lead. Lead: {lead.name} — Order: {order.order_number}',
                items=order_items,
                extra_data=extra_data,
                requested_by=user,
                lead=lead,
            )
            LeadActivity.objects.create(
                lead=lead, activity_type='status_change',
                description=f'Flow: Inventory approval {req_number} created for Order #{order.order_number}.',
                performed_by=user)
            result['executed'] = True
            result['message'] = f'Inventory approval {req_number} created — awaiting approval.'
        else:
            result['message'] = 'No linked order found to send to warehouse.'
        return result

    if flow.action == 'create_invoice' and flow.target_module == 'invoices':
        LeadActivity.objects.create(
            lead=lead, activity_type='status_change',
            description=f'Flow: Marked for invoicing.',
            performed_by=user)
        result['executed'] = True
        result['message'] = 'Marked for invoicing.'
        return result

    if flow.action == 'mark_dispatch' and flow.target_module == 'dispatch':
        from apps.orders.models import Order
        order = Order.objects.filter(external_id=f'lead-{lead.id}').first()
        if order:
            order.status = 'dispatched'
            order.save(update_fields=['status'])
            LeadActivity.objects.create(
                lead=lead, activity_type='status_change',
                description=f'Flow: Order #{order.order_number} marked for dispatch.',
                performed_by=user)
            result['executed'] = True
            result['message'] = f'Order #{order.order_number} marked for dispatch.'
        else:
            result['message'] = 'No linked order found.'
        return result

    result['message'] = f'Flow configured but action "{flow.action}" not yet implemented.'
    return result


class LeadDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET PUT PATCH DELETE /api/v1/crm/leads/<id>/"""
    serializer_class = LeadSerializer
    permission_classes = [IsStaffOrAbove]
    queryset = Lead.objects.all()

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        try:
            s, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
            ctx['lead_form_schema'] = s
        except Exception:
            ctx['lead_form_schema'] = None
        ctx['tenant'] = _crm_tenant()
        return ctx

    def retrieve(self, request, *args, **kwargs):
        return success_response(data=LeadSerializer(self.get_object()).data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        lead = self.get_object()
        old_status = lead.status
        serializer = self.get_serializer(lead, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        lead = serializer.save()
        flow_result = None
        if old_status != lead.status:
            LeadActivity.objects.create(
                lead=lead, activity_type='status_change',
                description=f'Status changed: {old_status} → {lead.status}',
                performed_by=request.user
            )
            flow_result = _execute_status_flow(lead, request.user)
        title = f'Status: {old_status} → {lead.status}' if old_status != lead.status else f'Lead updated: {lead.name}'
        log_module_history(request, 'crm', 'update', title, 'lead', str(lead.pk), {'old_status': old_status, 'new_status': lead.status})
        data = LeadSerializer(lead).data
        if flow_result:
            data['flow_result'] = flow_result
        return success_response(data=data, message='Lead updated.')

    def destroy(self, request, *args, **kwargs):
        lead = self.get_object()
        name, pk = lead.name, lead.pk
        log_module_history(request, 'crm', 'delete', f'Lead deleted: {name}', 'lead', str(pk))
        return super().destroy(request, *args, **kwargs)


class LeadFormSchemaView(APIView):
    """GET PUT /api/v1/crm/lead-form-schema/ — Get or update lead form schema. Admin-only for PUT."""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsStaffOrAbove()]
        return [IsTenantAdmin()]

    def _get_schema(self):
        from django.db import OperationalError
        try:
            schema, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
            return schema
        except OperationalError as e:
            if 'does not exist' in str(e).lower() or 'relation' in str(e).lower():
                raise ValueError('CRM schema table not found. Run: python manage.py migrate_schemas')
            raise

    def get(self, request):
        try:
            schema = self._get_schema()
        except ValueError as e:
            return error_response(str(e), http_status=status.HTTP_503_SERVICE_UNAVAILABLE)
        tenant = _crm_tenant()
        ser = LeadFormSchemaSerializer(schema)
        data = dict(ser.data)
        merged_defaults = merge_default_fields(schema, tenant=tenant)
        data['default_fields'] = merged_defaults
        data['custom_fields'] = schema.fields or []
        data['fields'] = merged_defaults + (schema.fields or [])
        data['phone_regex_presets'] = list(getattr(tenant, 'crm_phone_regex_presets', None) or []) if tenant else []
        return success_response(data=data)

    def put(self, request):
        from django.db import OperationalError
        try:
            schema = self._get_schema()
        except ValueError as e:
            return error_response(str(e), http_status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except OperationalError as e:
            return error_response(f'Database error: {e}', http_status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        try:
            data = dict(request.data) if hasattr(request.data, 'items') else {}
            custom_fields = data.get('custom_fields', data.get('fields', []))
            default_keys = {f['key'] for f in DEFAULT_LEAD_FIELDS}
            custom_fields = [f for f in custom_fields if f.get('key') not in default_keys]
            tenant = _crm_tenant()
            preset_ids = {str(p.get('id')) for p in (getattr(tenant, 'crm_phone_regex_presets', None) or []) if p.get('id')}

            payload = {'fields': custom_fields}
            overrides = data.get('default_field_overrides')
            if overrides is not None and isinstance(overrides, dict):
                clean = {}
                for k, v in overrides.items():
                    if k not in default_keys or not isinstance(v, dict):
                        continue
                    entry = {
                        x: y for x, y in v.items()
                        if x in ('order', 'required', 'label', 'phone_preset_id', 'pattern', 'phone_pattern')
                    }
                    if k == 'phone':
                        entry.pop('pattern', None)
                        entry.pop('phone_pattern', None)
                        pid = entry.get('phone_preset_id')
                        if pid is not None and str(pid).strip() and str(pid) not in preset_ids:
                            return error_response(
                                'Invalid phone format selection. Choose an option defined for this client.',
                                http_status=status.HTTP_400_BAD_REQUEST,
                            )
                    clean[k] = entry
                payload['default_field_overrides'] = clean
            serializer = LeadFormSchemaSerializer(schema, data=payload, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            schema.refresh_from_db()
            log_module_history(request, 'crm', 'settings_change', 'Lead form schema updated', 'lead_form_schema', '1', {'fields_count': len(custom_fields)})
            out = dict(LeadFormSchemaSerializer(schema).data)
            merged_defaults = merge_default_fields(schema, tenant=tenant)
            out['default_fields'] = merged_defaults
            out['custom_fields'] = schema.fields or []
            out['fields'] = merged_defaults + (schema.fields or [])
            out['phone_regex_presets'] = list(getattr(tenant, 'crm_phone_regex_presets', None) or []) if tenant else []
            return success_response(data=out, message='Lead form schema saved.')
        except Exception as e:
            return error_response(f'Could not save lead form schema: {str(e)}',
                                  http_status=status.HTTP_400_BAD_REQUEST)


class LeadFormPublicView(APIView):
    """GET /api/v1/crm/lead-form-public/ — Public schema for form (no auth). Returns default + custom fields."""
    permission_classes = [AllowAny]

    def get(self, request):
        schema, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
        tenant = _crm_tenant()
        form_fields = effective_form_fields(schema, exclude_display_only=True, tenant=tenant)
        # Source should be a dropdown from tenant's configured lead sources.
        source_opts = list(CustomLeadSource.objects.order_by('order', 'id').values_list('key', flat=True))
        if source_opts:
            for f in form_fields:
                if f.get('key') == 'source':
                    f['type'] = 'select'
                    f['options'] = source_opts
                    break
        tenant_info = {}
        if tenant:
            from apps.config.models import TenantConfig
            config = TenantConfig.objects.filter(tenant=tenant).first()
            if config:
                tenant_info = {
                    'company_name': config.company_name_override or tenant.name,
                    'logo': (config.logo.url if config.logo else None) or (tenant.logo.url if tenant.logo else None),
                }
        return success_response(data={'fields': form_fields, 'tenant_info': tenant_info})


class LeadSubmitView(APIView):
    """POST /api/v1/crm/lead-submit/ — Public form submit (no auth). Creates lead from form data."""
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def post(self, request):
        schema, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
        tenant = _crm_tenant()
        all_fields = effective_form_fields(schema, exclude_display_only=True, tenant=tenant)

        data = request.data if isinstance(request.data, dict) else dict(request.data)
        ok, err_msg, err_key = validate_required_from_schema(schema, data, tenant=tenant)
        if not ok:
            return error_response(err_msg, errors={err_key: [err_msg]} if err_key else None,
                                  http_status=status.HTTP_400_BAD_REQUEST)

        custom_data = {}
        name = ''
        email = ''
        phone = ''
        company = ''

        for field_def in all_fields:
            key = field_def.get('key')
            if not key or key in DISPLAY_ONLY_KEYS:
                continue
            val = data.get(key, '')
            if isinstance(val, list):
                val = val[0] if val else ''
            val = str(val).strip() if val else ''

            if key in LEAD_CORE_KEYS:
                if key == 'name':
                    name = val or name
                elif key == 'email':
                    email = val or email
                elif key == 'phone':
                    phone = val or phone
                elif key == 'company':
                    company = val or company
                # source, status: form submit uses fixed values
            else:
                if val:
                    custom_data[key] = val

        # Fallbacks for alternate keys
        if not name:
            name = (data.get('customer_name') or data.get('full_name') or data.get('client') or '').strip()
        if not email:
            email = (data.get('mail') or '').strip()
        if not phone:
            phone = (data.get('mobile') or data.get('contact') or data.get('contact_number') or '').strip()

        source_in = (data.get('source') or '').strip()
        valid_source = CustomLeadSource.objects.filter(key=source_in).exists() if source_in else False
        lead = Lead.objects.create(
            source=source_in if valid_source else LeadSource.FORM,
            status='new',
            name=name or 'Form submission',
            email=email or '',
            phone=phone or '',
            company=company or '',
            custom_data=custom_data,
        )
        _auto_link_customer(lead)
        auto_assign_lead_from_bulk_defaults(lead)
        log_module_history(request, 'crm', 'create', f'Lead from form: {lead.name}', 'lead', str(lead.pk))
        return success_response(
            data={'id': lead.id},
            message='Thank you! Your submission has been received.',
            http_status=status.HTTP_201_CREATED,
        )


class LeadFormTemplateView(APIView):
    """GET /api/v1/crm/lead-form-schema/template/ — Download Excel template."""
    permission_classes = [IsStaffOrAbove]

    def get(self, request):
        schema, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
        tenant = _crm_tenant()
        form_defaults = [f for f in merge_default_fields(schema, tenant=tenant) if f['key'] not in DISPLAY_ONLY_KEYS]
        fields = form_defaults + (schema.fields or [])
        fields = sorted(fields, key=lambda f: f.get('order', 999))
        seen = set()
        ordered = []
        for f in fields:
            k = f.get('key')
            if k and k not in seen:
                seen.add(k)
                ordered.append((k, f.get('label', k)))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Leads'
        for col, (_, label) in enumerate(ordered, 1):
            ws.cell(row=1, column=col, value=label)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=leads_template.xlsx'
        return response


class LeadFormImportView(APIView):
    """POST /api/v1/crm/lead-form-schema/import/ — Bulk upload Excel."""
    permission_classes = [IsStaffOrAbove]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return error_response('No file provided.', http_status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith(('.xlsx', '.xls')):
            return error_response('File must be Excel (.xlsx).', http_status=status.HTTP_400_BAD_REQUEST)

        schema, _ = LeadFormSchema.objects.get_or_create(pk=1, defaults={'fields': []})
        tenant = _crm_tenant()
        form_defaults = [f for f in merge_default_fields(schema, tenant=tenant) if f['key'] not in DISPLAY_ONLY_KEYS]
        all_field_keys = {f['key'] for f in form_defaults} | {f.get('key') for f in (schema.fields or []) if f.get('key')}

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=file, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in (header_row or [])]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
        except Exception as e:
            return error_response(f'Invalid Excel file: {str(e)}', http_status=status.HTTP_400_BAD_REQUEST)

        MAX_ROWS = 1000
        if len(rows) > MAX_ROWS:
            return error_response(f'Maximum {MAX_ROWS} rows allowed.', http_status=status.HTTP_400_BAD_REQUEST)

        core_keys = {'name', 'email', 'phone', 'company', 'source'}
        created = 0
        errors = []
        for i, row in enumerate(rows):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_num = i + 2
            data = {}
            for col_idx, val in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    key = headers[col_idx]
                    data[key] = str(val).strip() if val is not None else ''
            name = (data.get('name') or data.get('customer_name') or data.get('client') or '').strip()
            email = (data.get('email') or data.get('mail') or '').strip()
            phone = (data.get('phone') or data.get('mobile') or data.get('contact') or data.get('contact_number') or data.get('contact_numbe') or '').strip()
            company = (data.get('company') or '').strip()
            source_val = (data.get('source') or 'excel').strip() or 'excel'
            status_val = 'new'
            custom_data = {k: v for k, v in data.items() if k not in core_keys and v}
            if not name:
                name = email or phone or f'Row {row_num}'
            row_dict = {k: v for k, v in data.items()}
            ok_req, req_msg, _ = validate_required_from_schema(schema, row_dict, tenant=tenant)
            if not ok_req:
                errors.append({'row': row_num, 'message': req_msg})
                continue
            phone_field = next((f for f in merge_default_fields(schema, tenant=tenant) if f['key'] == 'phone'), None)
            if phone:
                pok, perr = validate_phone_value(phone, (phone_field or {}).get('pattern') if phone_field else None)
                if not pok:
                    errors.append({'row': row_num, 'message': perr})
                    continue
            try:
                lead = Lead.objects.create(
                    source=source_val,
                    status=status_val,
                    name=name,
                    email=email,
                    phone=phone,
                    company=company,
                    custom_data=custom_data,
                )
                _auto_link_customer(lead)
                auto_assign_lead_from_bulk_defaults(lead)
                created += 1
            except Exception as e:
                errors.append({'row': row_num, 'message': str(e)})

        if created:
            log_module_history(request, 'crm', 'create', f'Bulk import: {created} leads', 'lead', '', {'count': created})
        return success_response(data={'created': created, 'errors': errors}, message=f'Imported {created} leads.')


class LeadAssignmentConfigView(APIView):
    """
    GET PUT /api/v1/crm/lead-assignment/
    Legacy endpoint retained for backward compatibility.
    New lead assignment now uses CRM Bulk Assign defaults.
    """

    def get_permissions(self):
        return [IsTenantAdmin()]

    def get(self, request):
        cfg, _ = LeadAssignmentConfig.objects.get_or_create(pk=1, defaults={'strategy': 'off'})
        return success_response(data=LeadAssignmentConfigSerializer(cfg).data)

    def put(self, request):
        cfg, _ = LeadAssignmentConfig.objects.get_or_create(pk=1, defaults={'strategy': 'off'})
        serializer = LeadAssignmentConfigSerializer(cfg, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        log_module_history(request, 'crm', 'settings_change', 'Lead auto-assignment by source updated', 'lead_assignment', '1', {})
        return success_response(data=serializer.data, message='Assignment rules saved.')


class LeadAssignView(APIView):
    """POST /api/v1/crm/leads/<id>/assign/"""
    permission_classes = [IsStaffOrAbove]

    def post(self, request, pk):
        try:
            lead = Lead.objects.get(pk=pk)
        except Lead.DoesNotExist:
            return error_response('Lead not found.', http_status=404)

        user_id = request.data.get('user_id')
        from apps.users.models import User
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return error_response('User not found.')

        lead.assigned_to = user
        lead.save()
        LeadActivity.objects.create(
            lead=lead, activity_type='note',
            description=f'Lead assigned to {user.full_name}',
            performed_by=request.user
        )
        return success_response(data=LeadListSerializer(lead).data, message='Lead assigned.')


class CustomerListCreateView(generics.ListCreateAPIView):
    """GET POST /api/v1/crm/customers/"""
    serializer_class = CustomerSerializer
    permission_classes = [IsStaffOrAbove]

    def get_queryset(self):
        qs = Customer.objects.filter(is_active=True)
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(first_name__icontains=search) | qs.filter(
                last_name__icontains=search) | qs.filter(phone__icontains=search)
        return qs

    def list(self, request, *args, **kwargs):
        data = CustomerSerializer(self.get_queryset(), many=True).data
        return success_response(data={'customers': data, 'count': len(data)})

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        customer = serializer.save()
        return success_response(data=CustomerSerializer(customer).data,
                                message='Customer created.', http_status=status.HTTP_201_CREATED)


class CustomerDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET PUT PATCH DELETE /api/v1/crm/customers/<id>/"""
    serializer_class = CustomerSerializer
    permission_classes = [IsStaffOrAbove]
    queryset = Customer.objects.all()

    def retrieve(self, request, *args, **kwargs):
        return success_response(data=CustomerSerializer(self.get_object()).data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return success_response(data=serializer.data, message='Customer updated.')


class CustomerLeadsView(APIView):
    """GET /api/v1/crm/customers/<pk>/leads/ — all leads linked to a customer."""
    permission_classes = [IsStaffOrAbove]

    def get(self, request, pk):
        try:
            customer = Customer.objects.get(pk=pk)
        except Customer.DoesNotExist:
            return error_response('Customer not found.', http_status=status.HTTP_404_NOT_FOUND)
        leads = Lead.objects.filter(customer=customer).select_related('assigned_to')
        return success_response(data={
            'customer': CustomerSerializer(customer).data,
            'leads': LeadListSerializer(leads, many=True).data,
            'count': leads.count(),
        })


class LeadBulkAssignView(APIView):
    """
    POST /api/v1/crm/leads/bulk-assign/
    Bulk-assign selected (or all unassigned) leads using one of three strategies:
      round_robin — distribute evenly in rotation
      pool        — fill each employee's batch (pool_batch_size) then continue from remaining pool
      custom      — explicit count per employee  {employees: [{user_id, count}]}
    """
    permission_classes = [IsTenantAdmin]

    def post(self, request):
        from django.db import connection
        from apps.users.models import User
        from apps.hr.models import EmployeeProfile
        from apps.config.models import TenantConfig

        lead_ids          = request.data.get('lead_ids', [])
        filter_unassigned = request.data.get('filter_unassigned', False)
        assignment_type   = request.data.get('assignment_type', 'round_robin')
        employees_data    = request.data.get('employees', [])   # [{user_id, count?}]
        pool_batch_size   = int(request.data.get('pool_batch_size', 4))

        if not employees_data:
            return error_response('Provide at least one employee.')

        # Resolve User for each employee; backend expects user_id, but also accepts
        # employee_id (EmployeeProfile pk) when frontend sent wrong id
        resolved = []  # [(user, count), ...]; count used only for custom mode
        seen = set()
        for e in employees_data:
            uid = e.get('user_id')
            eid = e.get('employee_id') or e.get('id')
            count = int(e.get('count', 0))
            user = None
            if uid:
                try:
                    user = User.objects.get(pk=uid)
                except User.DoesNotExist:
                    pass
            if not user and eid:
                try:
                    emp = EmployeeProfile.objects.select_related('user').get(pk=eid)
                    if emp.user_id and emp.user_id not in seen:
                        user = emp.user
                except EmployeeProfile.DoesNotExist:
                    pass
            if user and user.id not in seen:
                seen.add(user.id)
                resolved.append((user, count))
        employees = [r[0] for r in resolved]
        if not employees:
            return error_response('No valid employees found. Ensure employees have linked user accounts.')

        # Resolve leads
        if lead_ids:
            leads = list(Lead.objects.filter(id__in=lead_ids))
        elif filter_unassigned:
            leads = list(Lead.objects.filter(assigned_to__isnull=True))
        else:
            return error_response('Provide lead_ids or set filter_unassigned=true.')

        if not leads:
            tenant = getattr(connection, 'tenant', None)
            if tenant:
                config, _ = TenantConfig.objects.get_or_create(tenant=tenant)
                config.crm_bulk_assign_defaults = {
                    'assignment_type': assignment_type,
                    'pool_batch_size': pool_batch_size,
                    'filter_unassigned': bool(filter_unassigned),
                    'employees': [{'user_id': u.id, 'count': c} for u, c in resolved],
                }
                config.save(update_fields=['crm_bulk_assign_defaults', 'updated_at'])
            return success_response(
                data={
                    'assigned': 0,
                    'total': 0,
                    'preferences_saved': True,
                },
                message=(
                    'No unassigned leads right now. Your strategy and selected employees were saved — '
                    'return here or run bulk assign from Leads when new leads are available.'
                ),
            )

        assigned = 0

        if assignment_type == 'round_robin':
            for i, lead in enumerate(leads):
                emp = employees[i % len(employees)]
                lead.assigned_to = emp
                lead.save(update_fields=['assigned_to'])
                LeadActivity.objects.create(lead=lead, activity_type='note',
                    description=f'Bulk assigned (round-robin) to {emp.full_name}',
                    performed_by=request.user)
                assigned += 1

        elif assignment_type == 'pool':
            idx = 0
            while idx < len(leads):
                for emp in employees:
                    for _ in range(pool_batch_size):
                        if idx >= len(leads):
                            break
                        lead = leads[idx]
                        lead.assigned_to = emp
                        lead.save(update_fields=['assigned_to'])
                        LeadActivity.objects.create(lead=lead, activity_type='note',
                            description=f'Bulk assigned (pool) to {emp.full_name}',
                            performed_by=request.user)
                        assigned += 1
                        idx += 1
                    if idx >= len(leads):
                        break

        elif assignment_type == 'custom':
            idx = 0
            for emp, count in resolved:
                if count <= 0:
                    continue
                for _ in range(count):
                    if idx >= len(leads):
                        break
                    lead = leads[idx]
                    lead.assigned_to = emp
                    lead.save(update_fields=['assigned_to'])
                    LeadActivity.objects.create(lead=lead, activity_type='note',
                        description=f'Bulk assigned (custom) to {emp.full_name}',
                        performed_by=request.user)
                    assigned += 1
                    idx += 1

        return success_response(
            data={'assigned': assigned, 'total': len(leads)},
            message=f'Successfully assigned {assigned} leads.',
        )


# ── Custom Lead Statuses ─────────────────────────────────────────────────────

DEFAULT_STATUSES = [
    {'key': 'new',           'label': 'New',           'color': '#3b82f6', 'order': 0},
    {'key': 'contacted',     'label': 'Contacted',     'color': '#f59e0b', 'order': 1},
    {'key': 'qualified',     'label': 'Qualified',     'color': '#10b981', 'order': 2},
    {'key': 'order_created', 'label': 'Order Created', 'color': '#8b5cf6', 'order': 3},
    {'key': 'lost',          'label': 'Lost',          'color': '#ef4444', 'order': 4},
]

DEFAULT_SOURCES = [
    {'key': 'meta',     'label': 'Meta / Facebook', 'order': 0},
    {'key': 'shopify',  'label': 'Shopify',         'order': 1},
    {'key': 'online',   'label': 'Online Order',    'order': 2},
    {'key': 'manual',   'label': 'Manual Entry',    'order': 3},
    {'key': 'whatsapp', 'label': 'WhatsApp',        'order': 4},
    {'key': 'referral', 'label': 'Referral',        'order': 5},
    {'key': 'form',     'label': 'Form Submission', 'order': 6},
    {'key': 'excel',    'label': 'Excel Import',    'order': 7},
]


def _seed_defaults():
    """Create default statuses/sources if none exist for this tenant."""
    if not CustomLeadStatus.objects.exists():
        for s in DEFAULT_STATUSES:
            CustomLeadStatus.objects.create(**s)
    if not CustomLeadSource.objects.exists():
        for s in DEFAULT_SOURCES:
            CustomLeadSource.objects.create(**s)


class CustomLeadStatusListCreateView(generics.ListCreateAPIView):
    serializer_class = CustomLeadStatusSerializer
    permission_classes = [IsStaffOrAbove]

    def get_queryset(self):
        return CustomLeadStatus.objects.all()

    def list(self, request, *args, **kwargs):
        _seed_defaults()
        data = CustomLeadStatusSerializer(self.get_queryset(), many=True).data
        return success_response(data={'statuses': data})

    def create(self, request, *args, **kwargs):
        self.permission_classes = [IsTenantAdmin]
        self.check_permissions(request)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return success_response(
            data=CustomLeadStatusSerializer(obj).data,
            message='Status created.', http_status=status.HTTP_201_CREATED)


class CustomLeadStatusDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CustomLeadStatusSerializer
    permission_classes = [IsTenantAdmin]
    queryset = CustomLeadStatus.objects.all()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return success_response(data=CustomLeadStatusSerializer(obj).data, message='Status updated.')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return success_response(message='Status deleted.')


# ── Custom Lead Sources ──────────────────────────────────────────────────────

class CustomLeadSourceListCreateView(generics.ListCreateAPIView):
    serializer_class = CustomLeadSourceSerializer
    permission_classes = [IsStaffOrAbove]

    def get_queryset(self):
        return CustomLeadSource.objects.all()

    def list(self, request, *args, **kwargs):
        _seed_defaults()
        data = CustomLeadSourceSerializer(self.get_queryset(), many=True).data
        return success_response(data={'sources': data})

    def create(self, request, *args, **kwargs):
        self.permission_classes = [IsTenantAdmin]
        self.check_permissions(request)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return success_response(
            data=CustomLeadSourceSerializer(obj).data,
            message='Source created.', http_status=status.HTTP_201_CREATED)


class CustomLeadSourceDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CustomLeadSourceSerializer
    permission_classes = [IsTenantAdmin]
    queryset = CustomLeadSource.objects.all()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return success_response(data=CustomLeadSourceSerializer(obj).data, message='Source updated.')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return success_response(message='Source deleted.')


# ── Status Flow Actions ──────────────────────────────────────────────────────

class StatusFlowActionListCreateView(generics.ListCreateAPIView):
    serializer_class = StatusFlowActionSerializer
    permission_classes = [IsStaffOrAbove]

    def get_queryset(self):
        return StatusFlowAction.objects.all()

    def list(self, request, *args, **kwargs):
        data = StatusFlowActionSerializer(self.get_queryset(), many=True).data
        return success_response(data={'flow_actions': data})

    def create(self, request, *args, **kwargs):
        self.permission_classes = [IsTenantAdmin]
        self.check_permissions(request)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return success_response(
            data=StatusFlowActionSerializer(obj).data,
            message='Flow action created.', http_status=status.HTTP_201_CREATED)


class StatusFlowActionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StatusFlowActionSerializer
    permission_classes = [IsTenantAdmin]
    queryset = StatusFlowAction.objects.all()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return success_response(data=StatusFlowActionSerializer(obj).data, message='Flow action updated.')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return success_response(message='Flow action deleted.')
