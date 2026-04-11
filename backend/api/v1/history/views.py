"""
History API — Client admin only.
List, export, and delete module history. Deletion creates permanent AuditSummary.
Export supports JSON (PDF) and Excel format.
"""
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response

from apps.config.models import ModuleHistory, AuditSummary, MODULE_HISTORY_CHOICES
from apps.core.responses import success_response, error_response
from apps.core.permissions import IsTenantAdmin


class ModuleHistoryListView(APIView):
    """GET /api/v1/history/ — List history (admin only). Filter by module, date."""
    permission_classes = [IsTenantAdmin]

    def get(self, request):
        qs = ModuleHistory.objects.filter(deleted_at__isnull=True).select_related('performed_by').order_by('-created_at')
        module = request.query_params.get('module')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if module:
            qs = qs.filter(module=module)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        data = []
        for h in qs[:500]:  # Limit for performance
            data.append({
                'id': h.id,
                'module': h.module,
                'action': h.action,
                'entity_type': h.entity_type,
                'entity_id': h.entity_id,
                'title': h.title,
                'details': h.details,
                'performed_by': h.performed_by.full_name if h.performed_by else h.performed_by_email or '—',
                'performed_by_email': h.performed_by_email or (h.performed_by.email if h.performed_by else ''),
                'created_at': h.created_at.isoformat() if h.created_at else None,
            })
        return success_response(data={'history': data, 'count': len(data)})


class ModuleHistoryDeleteView(APIView):
    """DELETE /api/v1/history/<pk>/ — Soft delete, create AuditSummary (permanent)."""
    permission_classes = [IsTenantAdmin]

    def delete(self, request, pk):
        try:
            h = ModuleHistory.objects.get(pk=pk, deleted_at__isnull=True)
        except ModuleHistory.DoesNotExist:
            return error_response('History record not found.', http_status=status.HTTP_404_NOT_FOUND)

        user = request.user
        user_name = user.full_name if user else ''
        user_email = user.email if user else ''
        summary_text = f'Admin deleted history: {h.module} — {h.title[:80]}'

        h.deleted_at = timezone.now()
        h.save()

        AuditSummary.objects.create(
            module=h.module,
            action='history_deleted',
            summary_text=summary_text,
            records_deleted=1,
            performed_by=user,
            performed_by_email=user_email,
        )

        return success_response(message='History deleted. Summary retained for audit.')


class ModuleHistoryExportView(APIView):
    """GET /api/v1/history/export/ — Export history. ?format=xlsx for Excel; otherwise JSON for PDF."""
    permission_classes = [IsTenantAdmin]

    def get(self, request):
        qs = ModuleHistory.objects.filter(deleted_at__isnull=True).select_related('performed_by').order_by('-created_at')
        module = request.query_params.get('module')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if module:
            qs = qs.filter(module=module)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        data = []
        for h in qs[:1000]:
            data.append({
                'id': h.id,
                'module': h.module,
                'action': h.action,
                'entity_type': h.entity_type,
                'entity_id': h.entity_id,
                'title': h.title,
                'details': h.details,
                'performed_by': h.performed_by.full_name if h.performed_by else h.performed_by_email or '—',
                'performed_by_email': h.performed_by_email or (h.performed_by.email if h.performed_by else ''),
                'created_at': h.created_at.isoformat() if h.created_at else None,
            })

        if request.query_params.get('format') == 'xlsx':
            from django.http import HttpResponse
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            wb = Workbook()
            ws = wb.active
            ws.title = 'Module History'

            headers = ['Date', 'Module', 'Action', 'Entity Type', 'Entity ID', 'Title', 'Performed By', 'Email']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            for row_idx, h in enumerate(data, 2):
                created = h['created_at'][:16].replace('T', ' ') if h.get('created_at') else '—'
                ws.cell(row=row_idx, column=1, value=created)
                ws.cell(row=row_idx, column=2, value=h.get('module', ''))
                ws.cell(row=row_idx, column=3, value=h.get('action', ''))
                ws.cell(row=row_idx, column=4, value=h.get('entity_type', ''))
                ws.cell(row=row_idx, column=5, value=h.get('entity_id', ''))
                ws.cell(row=row_idx, column=6, value=h.get('title', ''))
                ws.cell(row=row_idx, column=7, value=h.get('performed_by', ''))
                ws.cell(row=row_idx, column=8, value=h.get('performed_by_email', ''))

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col - 1]) + 2)

            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            fname = f'module_history_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{fname}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response

        return success_response(data={'history': data})
